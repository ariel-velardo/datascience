"""Ingestao auditada dos Indicadores de Trajetoria (coortes 2015-2020).

Le os seis XLSX originais (nunca os modifica), audita cada um isoladamente,
e so entao concatena numa camada analitica compacta em Parquet.

Saida:
    data/processed/trajetorias_2015_2020.parquet
    data/processed/auditoria_multicoorte.json   (todas as metricas apuradas)

Proveniencia: cada linha carrega NU_ANO_INGRESSO (coorte), ARQUIVO_ORIGEM e
MD5_ORIGEM, de modo que qualquer linha do painel e rastreavel ate o arquivo
publicado pelo Inep.

Uso:
    python src/constroi_painel_multicoorte.py
"""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

RAIZ = Path("data/raw/trajetoria")
DESTINO = Path("data/processed/trajetorias_2015_2020.parquet")
RELATORIO = Path("data/processed/auditoria_multicoorte.json")
COORTES = [2015, 2016, 2017, 2018, 2019, 2020]

LINHA_CABECALHO = 9

# Codigos CINE tem zero a esquerda (Char(2)/Char(7) no dicionario do Inep).
# Tipar como inteiro destroi os valores silenciosamente.
COLUNAS_TEXTO = {
    "NO_IES", "NO_CURSO", "CO_CINE_ROTULO", "NO_CINE_ROTULO",
    "CO_CINE_AREA_GERAL", "NO_CINE_AREA_GERAL",
}
COLUNAS_INT = {
    "CO_IES", "TP_CATEGORIA_ADMINISTRATIVA", "TP_ORGANIZACAO_ACADEMICA",
    "CO_CURSO", "TP_GRAU_ACADEMICO",
    "TP_MODALIDADE_ENSINO", "NU_ANO_INGRESSO", "NU_ANO_REFERENCIA",
    "NU_PRAZO_INTEGRALIZACAO", "NU_ANO_INTEGRALIZACAO",
    "NU_PRAZO_ACOMPANHAMENTO", "NU_ANO_MAXIMO_ACOMPANHAMENTO",
    "QT_INGRESSANTE", "QT_PERMANENCIA", "QT_CONCLUINTE", "QT_DESISTENCIA",
    "QT_FALECIDO",
}
COLUNAS_FLOAT = {"TAP", "TCA", "TDA", "TCAN", "TADA"}
# Geografia e estruturalmente nula em EAD.
COLUNAS_INT_NULAVEL = {"CO_REGIAO", "CO_UF", "CO_MUNICIPIO"}


def md5(caminho: Path, blocos: int = 1 << 20) -> str:
    h = hashlib.md5()
    with caminho.open("rb") as f:
        for bloco in iter(lambda: f.read(blocos), b""):
            h.update(bloco)
    return h.hexdigest()


def le_coorte(coorte: int):
    """Le um XLSX de coorte e devolve (dataframe, dicionario de auditoria)."""
    pasta = RAIZ / f"{coorte}_2024"
    xlsx = next(pasta.glob("*.xlsx"))

    aud = {
        "coorte": coorte,
        "arquivo": str(xlsx).replace("\\", "/"),
        "mb": round(xlsx.stat().st_size / 1e6, 2),
        "md5": md5(xlsx),
    }

    # Confere o md5 contra o arquivo publicado pelo Inep, quando presente.
    aud["md5_publicado_confere"] = None
    for txt in pasta.glob("md5*.txt"):
        conteudo = txt.read_text(encoding="utf-8", errors="replace")
        for linha in conteudo.splitlines():
            if xlsx.name in linha:
                aud["md5_publicado_confere"] = linha.split()[0] == aud["md5"]

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    aud["sheets"] = wb.sheetnames
    ws = wb[wb.sheetnames[0]]
    aud["max_row_excel"] = ws.max_row
    aud["max_col_excel"] = ws.max_column

    linhas = ws.iter_rows(min_row=1, values_only=True)
    preambulo = [next(linhas) for _ in range(LINHA_CABECALHO - 1)]
    aud["preambulo_linha5"] = str(preambulo[4][0])[:200]
    cabecalho = list(next(linhas))
    aud["colunas"] = [c for c in cabecalho if c is not None]
    aud["n_colunas"] = len(aud["colunas"])

    registros = []
    rodape = []
    for linha in linhas:
        if linha[0] is None:
            continue
        primeiro = str(linha[0]).strip()
        # A ultima linha do arquivo e "Fonte: Censo da Educacao Superior/Inep."
        if not primeiro.lstrip("-").isdigit():
            rodape.append(primeiro[:120])
            continue
        registros.append(linha[: len(cabecalho)])
    wb.close()

    aud["linhas_rodape_removidas"] = rodape
    df = pd.DataFrame(registros, columns=cabecalho)
    aud["linhas_dados"] = len(df)

    for col in df.columns:
        if col in COLUNAS_TEXTO:
            df[col] = df[col].astype("string")
        elif col in COLUNAS_INT_NULAVEL:
            df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
        elif col in COLUNAS_INT:
            df[col] = pd.to_numeric(df[col], errors="raise").astype("int64")
        elif col in COLUNAS_FLOAT:
            df[col] = pd.to_numeric(df[col], errors="raise").astype("float64")

    df["ARQUIVO_ORIGEM"] = xlsx.name
    df["MD5_ORIGEM"] = aud["md5"]
    return df, aud


def audita(df: pd.DataFrame, aud: dict) -> dict:
    """Aplica ao dataframe da coorte as verificacoes da auditoria anterior."""
    aud["anos_ingresso"] = sorted(df["NU_ANO_INGRESSO"].unique().tolist())
    aud["anos_referencia"] = sorted(df["NU_ANO_REFERENCIA"].unique().tolist())
    aud["n_cursos"] = int(df["CO_CURSO"].nunique())
    aud["n_ies"] = int(df["CO_IES"].nunique())
    aud["qt_ingressantes_coorte"] = int(
        df.groupby("CO_CURSO")["QT_INGRESSANTE"].first().sum()
    )

    chave = ["CO_CURSO", "NU_ANO_REFERENCIA"]
    aud["chave_unica"] = bool(not df.duplicated(chave).any())
    aud["linhas_integralmente_duplicadas"] = int(df.duplicated().sum())

    obs = df.groupby("CO_CURSO").size()
    aud["obs_por_curso"] = {
        str(k): int(v) for k, v in obs.value_counts().items()
    }
    aud["painel_balanceado"] = bool(obs.nunique() == 1)

    nulos = df.isna().sum()
    aud["colunas_com_nulos"] = {c: int(n) for c, n in nulos.items() if n > 0}
    ead = df["TP_MODALIDADE_ENSINO"] == 2
    aud["nulos_geograficos_sao_todos_ead"] = bool(
        df.loc[df["CO_REGIAO"].isna(), "TP_MODALIDADE_ENSINO"].eq(2).all()
        and df.loc[ead, "CO_REGIAO"].isna().all()
    )

    taxas = df[["TAP", "TCA", "TDA", "TCAN", "TADA"]]
    aud["valores_impossiveis"] = {
        "QT_INGRESSANTE<=0": int((df["QT_INGRESSANTE"] <= 0).sum()),
        "QT_negativas": int(
            (df[["QT_PERMANENCIA", "QT_CONCLUINTE",
                 "QT_DESISTENCIA", "QT_FALECIDO"]] < 0).any(axis=1).sum()
        ),
        "taxas_fora_0_100": int(
            ((taxas < 0) | (taxas > 100)).any(axis=1).sum()
        ),
    }
    aud["qt_ingressante_min"] = int(df["QT_INGRESSANTE"].min())

    # QT_INGRESSANTE deve ser constante dentro da trajetoria.
    aud["qt_ingressante_constante"] = bool(
        df.groupby("CO_CURSO")["QT_INGRESSANTE"].nunique().eq(1).all()
    )

    d = df.sort_values(["CO_CURSO", "NU_ANO_REFERENCIA"]).copy()
    g = d.groupby("CO_CURSO", sort=False)
    d["CUM_CONCLUINTE"] = g["QT_CONCLUINTE"].cumsum()
    d["CUM_DESISTENCIA"] = g["QT_DESISTENCIA"].cumsum()
    d["CUM_FALECIDO"] = g["QT_FALECIDO"].cumsum()

    # Identidade contabil acumulada.
    resid = (d["QT_PERMANENCIA"] + d["CUM_CONCLUINTE"] + d["CUM_DESISTENCIA"]
             + d["CUM_FALECIDO"] - d["QT_INGRESSANTE"])
    aud["identidade_contabil_ok"] = int((resid == 0).sum())
    aud["identidade_contabil_n"] = int(len(d))
    aud["identidade_contabil_resid_abs_max"] = int(resid.abs().max())

    # Denominador "vivo": QT_INGRESSANTE - falecidos acumulados.
    denom = d["QT_INGRESSANTE"] - d["CUM_FALECIDO"]
    tol = 0.005
    formulas = {
        "TAP": 100 * d["QT_PERMANENCIA"] / denom,
        "TCA": 100 * d["CUM_CONCLUINTE"] / denom,
        "TDA": 100 * d["CUM_DESISTENCIA"] / denom,
        "TCAN": 100 * d["QT_CONCLUINTE"] / denom,
        "TADA": 100 * d["QT_DESISTENCIA"] / denom,
    }
    aud["formulas_denominador_vivo"] = {
        k: int(((v - d[k]).abs() < tol).sum()) for k, v in formulas.items()
    }
    bruto = d["QT_INGRESSANTE"]
    aud["formulas_denominador_bruto"] = {
        "TDA": int(((100 * d["CUM_DESISTENCIA"] / bruto - d["TDA"]).abs()
                    < tol).sum()),
        "TADA": int(((100 * d["QT_DESISTENCIA"] / bruto - d["TADA"]).abs()
                     < tol).sum()),
    }

    soma = d["TAP"] + d["TCA"] + d["TDA"]
    aud["soma_TAP_TCA_TDA_igual_100"] = int(((soma - 100).abs() < 0.005).sum())
    aud["soma_TAP_TCA_TDA_desvio_max"] = float((soma - 100).abs().max())

    # Monotonicidade das acumuladas.
    aud["violacoes_TDA_decrescente"] = int(
        (d.groupby("CO_CURSO", sort=False)["TDA"].diff() < -1e-9).sum()
    )
    aud["violacoes_TAP_crescente"] = int(
        (d.groupby("CO_CURSO", sort=False)["TAP"].diff() > 1e-9).sum()
    )

    # Constancia das covariaveis dentro da trajetoria (snapshot 2024).
    covars = ["TP_CATEGORIA_ADMINISTRATIVA", "TP_ORGANIZACAO_ACADEMICA",
              "TP_GRAU_ACADEMICO", "TP_MODALIDADE_ENSINO",
              "CO_CINE_AREA_GERAL", "CO_IES", "NU_PRAZO_INTEGRALIZACAO"]
    aud["covariaveis_variantes_no_curso"] = {
        c: int((df.groupby("CO_CURSO")[c].nunique() > 1).sum())
        for c in covars
    }

    # Composicao por modalidade (snapshot 2024).
    ing = df.groupby("CO_CURSO").first()
    aud["modalidade"] = {}
    for k, v in ing["TP_MODALIDADE_ENSINO"].value_counts().items():
        sel = ing["TP_MODALIDADE_ENSINO"] == k
        aud["modalidade"][str(int(k))] = {
            "cursos": int(v),
            "ingressantes": int(ing.loc[sel, "QT_INGRESSANTE"].sum()),
        }
    return aud


def main() -> None:
    DESTINO.parent.mkdir(parents=True, exist_ok=True)
    auditorias = []
    partes = []
    for coorte in COORTES:
        print(f"[{coorte}] lendo...", flush=True)
        df, aud = le_coorte(coorte)
        print(f"[{coorte}] auditando {len(df):,} linhas...", flush=True)
        aud = audita(df, aud)
        auditorias.append(aud)
        partes.append(df)
        print(
            f"[{coorte}] cursos={aud['n_cursos']:,} ies={aud['n_ies']:,} "
            f"ingressantes={aud['qt_ingressantes_coorte']:,} "
            f"identidade={aud['identidade_contabil_ok']}/"
            f"{aud['identidade_contabil_n']}",
            flush=True,
        )

    # So concatena depois de auditar tudo, e so se os schemas forem identicos.
    colunas = [tuple(a["colunas"]) for a in auditorias]
    assert len(set(colunas)) == 1, "schemas divergentes entre coortes"

    painel = pd.concat(partes, ignore_index=True)
    painel["IDADE_TRAJETORIA"] = (
        painel["NU_ANO_REFERENCIA"] - painel["NU_ANO_INGRESSO"]
    ).astype("int16")

    # Acumulados, calculados dentro de (coorte, curso).
    painel = painel.sort_values(
        ["NU_ANO_INGRESSO", "CO_CURSO", "NU_ANO_REFERENCIA"]
    ).reset_index(drop=True)
    g = painel.groupby(["NU_ANO_INGRESSO", "CO_CURSO"], sort=False)
    painel["CUM_CONCLUINTE"] = g["QT_CONCLUINTE"].cumsum().astype("int64")
    painel["CUM_DESISTENCIA"] = g["QT_DESISTENCIA"].cumsum().astype("int64")
    painel["CUM_FALECIDO"] = g["QT_FALECIDO"].cumsum().astype("int64")

    # Populacao viva no inicio do ano t = permanencia ao fim de t-1;
    # para t=0 (ano de ingresso), os proprios ingressantes.
    painel["EM_RISCO_INICIO"] = (
        g["QT_PERMANENCIA"].shift(1).fillna(painel["QT_INGRESSANTE"])
        .astype("int64")
    )
    risco_vivo = painel["EM_RISCO_INICIO"] - painel["QT_FALECIDO"]
    painel["HAZARD_DESISTENCIA"] = (
        100 * painel["QT_DESISTENCIA"] / risco_vivo
    ).where(risco_vivo > 0)

    tabela = pa.Table.from_pandas(painel, preserve_index=False)
    pq.write_table(tabela, DESTINO, compression="zstd")

    resumo = {
        "linhas": int(len(painel)),
        "coortes": COORTES,
        "trajetorias_coorte_curso": int(
            painel.groupby(["NU_ANO_INGRESSO", "CO_CURSO"]).ngroups
        ),
        "cursos_distintos_no_painel": int(painel["CO_CURSO"].nunique()),
        "ies_distintas_no_painel": int(painel["CO_IES"].nunique()),
        "anos_referencia": sorted(
            painel["NU_ANO_REFERENCIA"].unique().tolist()),
        "idades_trajetoria": sorted(
            painel["IDADE_TRAJETORIA"].unique().tolist()),
        "parquet_mb": round(DESTINO.stat().st_size / 1e6, 2),
    }
    RELATORIO.write_text(
        json.dumps({"resumo": resumo, "por_coorte": auditorias},
                   ensure_ascii=False, indent=1),
        encoding="utf-8",
    )
    print(json.dumps(resumo, ensure_ascii=False, indent=1))
    print(f"\nParquet: {DESTINO}  ({resumo['parquet_mb']} MB)")
    print(f"Auditoria: {RELATORIO}")


if __name__ == "__main__":
    main()
