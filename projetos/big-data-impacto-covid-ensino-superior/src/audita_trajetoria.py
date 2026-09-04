from pathlib import Path
from openpyxl import load_workbook

arquivo = Path(
    "data/raw/trajetoria/2015_2024/"
    "indicadores_trajetoria_educacao_superior_2015_2024.xlsx"
)

print("=" * 80)
print("AUDITORIA - INDICADORES DE TRAJETÓRIA")
print("=" * 80)

print(f"\nArquivo: {arquivo}")
print(f"Tamanho: {arquivo.stat().st_size / 1024**2:.2f} MB")

wb = load_workbook(
    arquivo,
    read_only=True,
    data_only=True
)

ws = wb["INDICADORES_TRAJETORIA"]

print(f"\nDimensão informada pelo Excel:")
print(f"Linhas: {ws.max_row:,}")
print(f"Colunas: {ws.max_column:,}")

# Analisa as primeiras 60 linhas para localizar automaticamente
# a linha mais provável de cabeçalho.
primeiras = list(
    ws.iter_rows(
        min_row=1,
        max_row=min(60, ws.max_row),
        values_only=True
    )
)

contagens = [
    sum(valor is not None for valor in linha)
    for linha in primeiras
]

header_idx = max(
    range(len(contagens)),
    key=lambda i: contagens[i]
)

header_row = header_idx + 1
cabecalho = primeiras[header_idx]

print("\n" + "=" * 80)
print("CANDIDATO A CABEÇALHO")
print("=" * 80)

print(
    f"\nLinha candidata: {header_row} "
    f"({contagens[header_idx]} células preenchidas)"
)

for i, valor in enumerate(cabecalho, start=1):
    print(f"{i:02d}. {valor}")

print("\n" + "=" * 80)
print("3 PRIMEIROS REGISTROS APÓS O CABEÇALHO")
print("=" * 80)

for numero_linha, linha in enumerate(
    ws.iter_rows(
        min_row=header_row + 1,
        max_row=header_row + 3,
        values_only=True
    ),
    start=header_row + 1
):
    print(f"\n--- Linha {numero_linha} ---")
    for coluna, valor in zip(cabecalho, linha):
        print(f"{coluna}: {valor}")

wb.close()
